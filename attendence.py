import datetime as dt
import openpyxl as ox
from openpyxl.styles import Font,Alignment
import os
import pandas as pd
import streamlit as sl
import os
import sys
base_path = getattr(sys, "_MEIPASS", os.path.abspath(os.path.dirname(__file__)))
current_date = dt.date.today()
current_time = dt.datetime.now().time()
hour=current_time.hour
minute=current_time.minute
current_time=(f'{hour}:{minute}')
current_day=current_date.weekday()
days=['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
current_day=days[current_day]
if not os.path.exists(f'Data/Attandance_list/Present_list_{current_date}.xlsx'):
    xl = ox.Workbook() 
    xl_act = xl.active
    heading_font=Font(bold=True,size=16,color='33022c')
    alignment = Alignment(horizontal="center")
    for row in xl_act.iter_rows():  # Loop through all rows in the sheet
        for cell in row:
            cell.alignment = alignment
    xl_act.merge_cells('A1:C1')
    xl_act['A1']=f'Present_list : {current_date}'
    xl_act['A2']='Name'
    xl_act['B2']='Time'
    xl_act['A1'].font=heading_font
    xl_act['A2'].font=heading_font
    xl_act['B2'].font=heading_font
    xl.save(f'Data/Attandance_list/Present_list_{current_date}.xlsx')
def present(name):
    wxl=ox.load_workbook(f'Data/Attandance_list/Present_list_{current_date}.xlsx')
    wxl_act=wxl.active
    wxl_act.append([f'{name}',f'{current_time}'])
    wxl.save(f'Data/Attandance_list/Present_list_{current_date}.xlsx')